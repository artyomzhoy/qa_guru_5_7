import os

from openpyxl import load_workbook

from os_path.os_path_scripts import PROJECT_ROOT_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    xlsx_path = os.path.join(PROJECT_ROOT_PATH, '..', 'resources', 'file_example_XLSX_50.xlsx')

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    value = sheet.cell(row=3, column=2).value
    assert value == 'Mara'
